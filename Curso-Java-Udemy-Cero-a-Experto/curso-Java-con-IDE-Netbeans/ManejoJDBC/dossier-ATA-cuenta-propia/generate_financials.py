#!/usr/bin/env python3
"""Generate realistic 3-year financial workbook for ATA business plan."""

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Styles ---
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF")
section_fill = PatternFill("solid", fgColor="D6E3F0")
money_format = '#,##0.00'
thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin


def autosize(ws, min_width=12, max_width=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


# Fixed monthly operating costs (EUR) — prudent / defendable
RETA = 89.00          # tarifa plana 80 + MEI approx
GESTORIA = 60.00
SAAS_CLOUD = 110.00
INTERNET_UTILS = 50.00
MANUTENCION = 650.00  # >= IPREM ~600
RC_PROF = 20.00
FIXED_MONTHLY = RETA + GESTORIA + SAAS_CLOUD + INTERNET_UTILS + MANUTENCION + RC_PROF  # 979

# CapEx already owned (placeholders — replace with invoice amounts)
CAPEX = {
    "Apple MacBook Pro (workstation)": 2_400.00,
    "2x LG UltraGear 27\" monitors": 450.00,
    "Ergonomic chair": 650.00,
    "Height-adjustable desk": 280.00,
}
CAPEX_TOTAL = sum(CAPEX.values())  # 3_780
CASH_SEED = 10_000.00
FAMILY_LOAN = 10_000.00

# Base scenario — monthly revenue year 1 (prudent ramp)
BASE_REV = [150, 200, 300, 700, 900, 1_000, 1_300, 1_500, 1_600, 1_800, 2_000, 2_200]
# Pessimistic: slower acquisition
PESS_REV = [0, 100, 150, 300, 450, 550, 700, 850, 950, 1_100, 1_200, 1_350]
# Optimistic: still realistic for B2B freelance
OPT_REV = [250, 400, 600, 1_000, 1_200, 1_400, 1_700, 2_000, 2_200, 2_400, 2_600, 2_800]

MONTHS = [f"M{i}" for i in range(1, 13)]

# ===================== Sheet 1: Assumptions =====================
ws = wb.active
ws.title = "Assumptions"
ws["A1"] = "FINANCIAL ASSUMPTIONS — Business Plan ATA (Self-Employment)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Promoter: Ricardo [Last Name] | Sector: CNAE 6202 | Location: Tarragona (Catalonia)"
ws["A3"] = "Currency: EUR | Horizon: 3 years | Scenario focus: Base (prudent)"

rows = [
    ("Parameter", "Value", "Notes"),
    ("Seed cash (family loan 0%)", CASH_SEED, "Working capital / runway only — not for CapEx"),
    ("CapEx already executed", CAPEX_TOTAL, "Replace with invoice totals"),
    ("RETA (flat rate + MEI)", RETA, "Approx. 80 + MEI 2026"),
    ("Accounting firm", GESTORIA, "Monthly"),
    ("SaaS / cloud / VPN", SAAS_CLOUD, "AWS + tools"),
    ("Internet / utilities share", INTERNET_UTILS, "Home office allocation"),
    ("Living costs / household", MANUTENCION, "At or above IPREM (~600)"),
    ("Professional liability (monthly avg)", RC_PROF, "RC profesional"),
    ("Total fixed monthly burn", FIXED_MONTHLY, "Base burn rate"),
    ("Runway at zero revenue (months)", round(CASH_SEED / FIXED_MONTHLY, 1), "Cash / burn"),
    ("IPREM monthly (ref. 2026)", 600.00, "Subsistence reference"),
    ("Year 2 revenue growth vs Y1", "20%", "Moderate"),
    ("Year 3 revenue growth vs Y2", "18%", "Moderate"),
]
for i, row in enumerate(rows, start=5):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)
        ws.cell(row=i, column=j).border = thin
        if i == 5:
            ws.cell(row=i, column=j).fill = header_fill
            ws.cell(row=i, column=j).font = header_font
        elif j == 2 and isinstance(val, float):
            ws.cell(row=i, column=j).number_format = money_format
autosize(ws)

# ===================== Sheet 2: Opening Balance =====================
ws = wb.create_sheet("Opening_Balance")
ws["A1"] = "OPENING BALANCE SHEET (Day 0)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Asset-light model: CapEx already owned; 10,000 EUR cash as working capital"

ws["A4"] = "ASSETS"
ws["A4"].fill = section_fill
ws["A4"].font = Font(bold=True)
headers = ["Item", "Amount (EUR)", "Evidence"]
for j, h in enumerate(headers, 1):
    ws.cell(row=5, column=j, value=h)
style_header(ws, 5, 3)

r = 6
ws.cell(row=r, column=1, value="Cash and cash equivalents (seed)")
ws.cell(row=r, column=2, value=CASH_SEED)
ws.cell(row=r, column=3, value="Bank transfer + Modelo 600 + loan contract")
r = 7
ws.cell(row=r, column=1, value="Property, plant & equipment (home office)")
ws.cell(row=r, column=2, value=CAPEX_TOTAL)
ws.cell(row=r, column=3, value="Invoices / alternative proof — Annex CapEx")
r = 8
for name, amount in CAPEX.items():
    r += 1
    ws.cell(row=r, column=1, value=f"   - {name}")
    ws.cell(row=r, column=2, value=amount)
    ws.cell(row=r, column=3, value="Adjust to invoice")

r_total_assets = r + 2
ws.cell(row=r_total_assets, column=1, value="TOTAL ASSETS")
ws.cell(row=r_total_assets, column=1).font = Font(bold=True)
ws.cell(row=r_total_assets, column=2, value=CASH_SEED + CAPEX_TOTAL)
ws.cell(row=r_total_assets, column=2).font = Font(bold=True)

r = r_total_assets + 2
ws.cell(row=r, column=1, value="EQUITY & LIABILITIES")
ws.cell(row=r, column=1).fill = section_fill
ws.cell(row=r, column=1).font = Font(bold=True)

r += 1
for j, h in enumerate(headers, 1):
    ws.cell(row=r, column=j, value=h)
style_header(ws, r, 3)

r += 1
ws.cell(row=r, column=1, value="Family loan (0% interest) — Renzo [Last Name]")
ws.cell(row=r, column=2, value=FAMILY_LOAN)
ws.cell(row=r, column=3, value="Private loan contract + Modelo 600")
r += 1
ws.cell(row=r, column=1, value="Owner contribution in kind (equipment)")
ws.cell(row=r, column=2, value=CAPEX_TOTAL)
ws.cell(row=r, column=3, value="Equipment already owned by promoter")
r += 2
ws.cell(row=r, column=1, value="TOTAL EQUITY & LIABILITIES")
ws.cell(row=r, column=1).font = Font(bold=True)
ws.cell(row=r, column=2, value=FAMILY_LOAN + CAPEX_TOTAL)
ws.cell(row=r, column=2).font = Font(bold=True)

for row in ws.iter_rows(min_row=6, max_row=r, min_col=2, max_col=2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = money_format
            cell.border = thin
autosize(ws)

# ===================== Sheet 3: Year1 Monthly =====================
ws = wb.create_sheet("Year1_Monthly_Base")
ws["A1"] = "YEAR 1 — MONTHLY P&L AND CASH FLOW (BASE SCENARIO)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = f"Fixed monthly costs: {FIXED_MONTHLY:.2f} EUR | Opening cash: {CASH_SEED:.2f} EUR"

headers = ["Concept"] + MONTHS + ["YEAR TOTAL"]
for j, h in enumerate(headers, 1):
    ws.cell(row=4, column=j, value=h)
style_header(ws, 4, len(headers))

ws.cell(row=5, column=1, value="Revenue (services)")
for i, rev in enumerate(BASE_REV, start=2):
    ws.cell(row=5, column=i, value=rev)
ws.cell(row=5, column=14, value=sum(BASE_REV))

cost_rows = [
    (6, "RETA (SS)", RETA),
    (7, "Accounting firm", GESTORIA),
    (8, "SaaS / cloud / VPN", SAAS_CLOUD),
    (9, "Internet / utilities share", INTERNET_UTILS),
    (10, "Living / household contribution", MANUTENCION),
    (11, "Professional liability (avg)", RC_PROF),
]
for row_i, label, amount in cost_rows:
    ws.cell(row=row_i, column=1, value=label)
    for c in range(2, 14):
        ws.cell(row=row_i, column=c, value=amount)
    ws.cell(row=row_i, column=14, value=amount * 12)

ws.cell(row=12, column=1, value="Total operating costs")
ws.cell(row=12, column=1).font = Font(bold=True)
for c in range(2, 14):
    ws.cell(row=12, column=c, value=FIXED_MONTHLY)
ws.cell(row=12, column=14, value=FIXED_MONTHLY * 12)

ws.cell(row=13, column=1, value="Operating result (Rev - Costs)")
ws.cell(row=13, column=1).font = Font(bold=True)
monthly_result = []
for i, rev in enumerate(BASE_REV):
    result = rev - FIXED_MONTHLY
    monthly_result.append(result)
    ws.cell(row=13, column=i + 2, value=result)
ws.cell(row=13, column=14, value=sum(monthly_result))

ws.cell(row=15, column=1, value="Opening cash")
ws.cell(row=16, column=1, value="Net cash movement")
ws.cell(row=17, column=1, value="Closing cash")
ws.cell(row=17, column=1).font = Font(bold=True)

cash = CASH_SEED
closing = []
for i, result in enumerate(monthly_result):
    open_c = cash
    ws.cell(row=15, column=i + 2, value=round(open_c, 2))
    ws.cell(row=16, column=i + 2, value=round(result, 2))
    cash = open_c + result
    closing.append(cash)
    ws.cell(row=17, column=i + 2, value=round(cash, 2))
ws.cell(row=15, column=14, value=CASH_SEED)
ws.cell(row=16, column=14, value=sum(monthly_result))
ws.cell(row=17, column=14, value=round(closing[-1], 2))

ws.cell(row=19, column=1, value="Break-even note")
ws.cell(row=19, column=2, value="Technical break-even when monthly revenue >= fixed costs (~979 EUR). In base scenario: around Month 5-6.")

for row in ws.iter_rows(min_row=5, max_row=17, min_col=2, max_col=14):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = money_format
            cell.border = thin

# Chart data block for Excel chart
ws.cell(row=21, column=1, value="Chart data")
ws.cell(row=22, column=1, value="Month")
ws.cell(row=23, column=1, value="Revenue")
ws.cell(row=24, column=1, value="Fixed costs")
for i, m in enumerate(MONTHS):
    ws.cell(row=22, column=i + 2, value=m)
    ws.cell(row=23, column=i + 2, value=BASE_REV[i])
    ws.cell(row=24, column=i + 2, value=FIXED_MONTHLY)

chart = LineChart()
chart.title = "Revenue vs Fixed Costs — Year 1 (Base)"
chart.style = 10
chart.y_axis.title = "EUR"
chart.x_axis.title = "Month"
chart.height = 10
chart.width = 18
data = Reference(ws, min_col=1, min_row=23, max_col=13, max_row=24)
cats = Reference(ws, min_col=2, max_col=13, min_row=22)
chart.add_data(data, from_rows=True, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "A26")
autosize(ws, min_width=10, max_width=22)

# ===================== Sheet 4: Scenarios =====================
ws = wb.create_sheet("Scenarios_Y1")
ws["A1"] = "YEAR 1 — SCENARIO COMPARISON (Pessimistic / Base / Optimistic)"
ws["A1"].font = Font(bold=True, size=14)

ws.cell(row=3, column=1, value="Metric")
ws.cell(row=3, column=2, value="Pessimistic")
ws.cell(row=3, column=3, value="Base")
ws.cell(row=3, column=4, value="Optimistic")
style_header(ws, 3, 4)


def year_metrics(revs):
    total_rev = sum(revs)
    total_cost = FIXED_MONTHLY * 12
    result = total_rev - total_cost
    cash = CASH_SEED
    min_cash = cash
    for rev in revs:
        cash += rev - FIXED_MONTHLY
        min_cash = min(min_cash, cash)
    return total_rev, total_cost, result, cash, min_cash


metrics = {
    "Pessimistic": year_metrics(PESS_REV),
    "Base": year_metrics(BASE_REV),
    "Optimistic": year_metrics(OPT_REV),
}

labels = [
    "Total revenue Y1",
    "Total fixed costs Y1",
    "Operating result Y1",
    "Closing cash Y1",
    "Minimum cash during Y1",
]
for i, label in enumerate(labels):
    ws.cell(row=4 + i, column=1, value=label)
    ws.cell(row=4 + i, column=2, value=metrics["Pessimistic"][i])
    ws.cell(row=4 + i, column=3, value=metrics["Base"][i])
    ws.cell(row=4 + i, column=4, value=metrics["Optimistic"][i])

ws.cell(row=10, column=1, value="Monthly revenue detail")
ws.cell(row=10, column=1).font = Font(bold=True)
for j, h in enumerate(["Month", "Pessimistic", "Base", "Optimistic"], 1):
    ws.cell(row=11, column=j, value=h)
style_header(ws, 11, 4)
for i in range(12):
    ws.cell(row=12 + i, column=1, value=MONTHS[i])
    ws.cell(row=12 + i, column=2, value=PESS_REV[i])
    ws.cell(row=12 + i, column=3, value=BASE_REV[i])
    ws.cell(row=12 + i, column=4, value=OPT_REV[i])

ws.cell(row=25, column=1, value="Stress test conclusion")
ws.cell(
    row=26,
    column=1,
    value=(
        "Even in the pessimistic scenario, minimum cash stays positive thanks to the 10,000 EUR "
        "seed. Living costs remain covered at or above IPREM throughout Year 1."
    ),
)

for row in ws.iter_rows(min_row=4, max_row=23, min_col=2, max_col=4):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = money_format
autosize(ws)

# ===================== Sheet 5: 3-Year =====================
ws = wb.create_sheet("Three_Year_Summary")
ws["A1"] = "THREE-YEAR P&L SUMMARY (BASE SCENARIO)"
ws["A1"].font = Font(bold=True, size=14)

y1_rev = sum(BASE_REV)
y1_cost = FIXED_MONTHLY * 12
y1_result = y1_rev - y1_cost
y1_close = CASH_SEED + y1_result

# Year 2-3: moderate growth; fixed costs rise slightly (RETA may leave flat rate)
y2_rev = round(y1_rev * 1.20, 2)
y3_rev = round(y2_rev * 1.18, 2)
# After year 1, assume higher SS contribution if income rises — still prudent
y2_fixed_monthly = 1_150.00  # higher RETA + same living + tools
y3_fixed_monthly = 1_250.00
y2_cost = y2_fixed_monthly * 12
y3_cost = y3_fixed_monthly * 12
y2_result = y2_rev - y2_cost
y3_result = y3_rev - y3_cost
y2_close = y1_close + y2_result
y3_close = y2_close + y3_result

headers = ["Concept", "Year 1", "Year 2", "Year 3"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 4)

data_rows = [
    ("Revenue", y1_rev, y2_rev, y3_rev),
    ("Operating costs (incl. living allocation)", y1_cost, y2_cost, y3_cost),
    ("Operating result", y1_result, y2_result, y3_result),
    ("Closing cash (cumulative)", y1_close, y2_close, y3_close),
    ("Avg monthly revenue", round(y1_rev / 12, 2), round(y2_rev / 12, 2), round(y3_rev / 12, 2)),
    ("Avg monthly fixed costs", FIXED_MONTHLY, y2_fixed_monthly, y3_fixed_monthly),
]
for i, row in enumerate(data_rows, start=4):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)
        if j > 1 and isinstance(val, float):
            ws.cell(row=i, column=j).number_format = money_format

ws.cell(row=11, column=1, value="Notes")
ws["A12"] = (
    "Years 2-3 assume gradual growth after reputation building. "
    "Fixed costs increase mainly due to higher RETA contribution when leaving the flat rate "
    "and slightly higher tooling. No aggressive revenue jumps. Figures are estimates for "
    "viability assessment, not guaranteed outcomes."
)
ws.merge_cells("A12:D14")
ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")

ws["A16"] = "Billable hours logic (illustrative, Base Y1 average)"
ws["A17"] = "If average rate = 45 EUR/h (mid-market remote), Year 1 revenue implies approx.:"
ws["A18"] = round(y1_rev / 45, 1)
ws["B18"] = "billable hours in the full year (~ "
ws["C18"] = round((y1_rev / 45) / 12, 1)
ws["D18"] = "h/month on average — conservative for a solo consultant)"

autosize(ws)

# ===================== Sheet 6: CapEx inventory =====================
ws = wb.create_sheet("CapEx_Inventory")
ws["A1"] = "CAPEX INVENTORY — ALREADY EXECUTED (HOME OFFICE)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Replace estimated amounts with invoice figures before submission to ATA."

headers = ["Asset", "Estimated EUR", "Owner", "Invoice attached (Y/N)", "Notes"]
for j, h in enumerate(headers, 1):
    ws.cell(row=4, column=j, value=h)
style_header(ws, 4, 5)

r = 5
for name, amount in CAPEX.items():
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=amount)
    ws.cell(row=r, column=2).number_format = money_format
    ws.cell(row=r, column=3, value="Ricardo [Last Name] / or Renzo + assignment")
    ws.cell(row=r, column=4, value="[ ]")
    ws.cell(row=r, column=5, value="Insert invoice or bank proof")
    r += 1

ws.cell(row=r, column=1, value="TOTAL CAPEX")
ws.cell(row=r, column=1).font = Font(bold=True)
ws.cell(row=r, column=2, value=CAPEX_TOTAL)
ws.cell(row=r, column=2).number_format = money_format
ws.cell(row=r, column=2).font = Font(bold=True)

ws.cell(row=r + 2, column=1, value="Important")
ws[f"A{r+3}"] = (
    "The 10,000 EUR seed is NOT used to purchase this equipment. "
    "Equipment is already available; cash is reserved exclusively for operating runway."
)
autosize(ws)

out = "Plan_Financiero_3_Anos_ATA.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"Fixed monthly burn: {FIXED_MONTHLY}")
print(f"Y1 base revenue: {sum(BASE_REV)}")
print(f"Y1 base closing cash: {round(CASH_SEED + sum(BASE_REV) - FIXED_MONTHLY * 12, 2)}")
print(f"Pess min cash check: {year_metrics(PESS_REV)[4]:.2f}")
